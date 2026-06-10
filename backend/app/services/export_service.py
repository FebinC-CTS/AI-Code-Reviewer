import io
import logging
from typing import List
from collections import defaultdict

from app.models.schemas import ReviewIssue, SeverityLevel

logger = logging.getLogger(__name__)

SEVERITY_ORDER = {SeverityLevel.HIGH: 0, SeverityLevel.MEDIUM: 1, SeverityLevel.LOW: 2}
SEVERITY_EMOJI = {SeverityLevel.HIGH: "🔴", SeverityLevel.MEDIUM: "🟡", SeverityLevel.LOW: "🟢"}


class ExportService:

    def export_excel(self, issues: List[ReviewIssue]) -> bytes:
        """Export issues to Excel (.xlsx) format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Code Review Issues"

        headers = ["File", "Issue", "Severity", "Explanation", "Fix", "Recommendation", "Insights", "Source"]

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        severity_fills = {
            "High": PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid"),
            "Medium": PatternFill(start_color="FFE0A3", end_color="FFE0A3", fill_type="solid"),
            "Low": PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid"),
        }

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 30

        # Sort issues by severity
        sorted_issues = sorted(issues, key=lambda x: SEVERITY_ORDER.get(x.severity, 3))

        for row_idx, issue in enumerate(sorted_issues, start=2):
            row_data = [
                issue.file,
                issue.issue,
                issue.severity.value,
                issue.explanation,
                issue.fix,
                issue.recommendation,
                issue.insights,
                issue.source or "ai",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

                # Color severity column
                if col_idx == 3:
                    sev_fill = severity_fills.get(issue.severity.value)
                    if sev_fill:
                        cell.fill = sev_fill
                        cell.font = Font(bold=True)

        # Auto-size columns with sensible max widths
        col_max_widths = [50, 40, 12, 60, 60, 60, 60, 10]
        for col_idx, (header, max_width) in enumerate(zip(headers, col_max_widths), start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, max(15, len(header) + 4))

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary["A1"] = "Code Review Summary"
        ws_summary["A1"].font = Font(bold=True, size=14)

        high_count = sum(1 for i in issues if i.severity == SeverityLevel.HIGH)
        med_count = sum(1 for i in issues if i.severity == SeverityLevel.MEDIUM)
        low_count = sum(1 for i in issues if i.severity == SeverityLevel.LOW)
        unique_files = len(set(i.file for i in issues))

        summary_data = [
            ("Total Issues", len(issues)),
            ("High Severity", high_count),
            ("Medium Severity", med_count),
            ("Low Severity", low_count),
            ("Files with Issues", unique_files),
        ]
        for row_idx, (label, value) in enumerate(summary_data, start=3):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def export_markdown(self, issues: List[ReviewIssue]) -> str:
        """Export issues to Markdown format."""
        if not issues:
            return "# Code Review Report\n\nNo issues found.\n"

        high_count = sum(1 for i in issues if i.severity == SeverityLevel.HIGH)
        med_count = sum(1 for i in issues if i.severity == SeverityLevel.MEDIUM)
        low_count = sum(1 for i in issues if i.severity == SeverityLevel.LOW)
        unique_files = len(set(i.file for i in issues))

        lines = [
            "# AI Code Review Report",
            "",
            "## Summary",
            "",
            f"| Metric | Count |",
            f"|--------|-------|",
            f"| Total Issues | {len(issues)} |",
            f"| 🔴 High Severity | {high_count} |",
            f"| 🟡 Medium Severity | {med_count} |",
            f"| 🟢 Low Severity | {low_count} |",
            f"| Files with Issues | {unique_files} |",
            "",
            "---",
            "",
        ]

        # Group by file
        by_file: dict[str, List[ReviewIssue]] = defaultdict(list)
        for issue in issues:
            by_file[issue.file].append(issue)

        for file_path in sorted(by_file.keys()):
            file_issues = sorted(by_file[file_path], key=lambda x: SEVERITY_ORDER.get(x.severity, 3))
            lines.append(f"## `{file_path}`")
            lines.append("")

            for idx, issue in enumerate(file_issues, start=1):
                emoji = SEVERITY_EMOJI.get(issue.severity, "⚪")
                lines.append(f"### {idx}. {emoji} [{issue.severity.value}] {issue.issue}")
                lines.append("")
                lines.append(f"**Explanation:** {issue.explanation}")
                lines.append("")

                if issue.fix:
                    lines.append("**Fix:**")
                    lines.append("")
                    # Detect if fix looks like code
                    if "\n" in issue.fix or issue.fix.strip().startswith(("def ", "function", "const ", "let ", "var ", "class ", "import ", "from ")):
                        lang = _guess_lang(file_path)
                        lines.append(f"```{lang}")
                        lines.append(issue.fix)
                        lines.append("```")
                    else:
                        lines.append(issue.fix)
                    lines.append("")

                if issue.recommendation:
                    lines.append(f"**Recommendation:** {issue.recommendation}")
                    lines.append("")

                if issue.insights:
                    lines.append(f"**Insights:** {issue.insights}")
                    lines.append("")

                source_label = "🤖 AI Analysis" if issue.source == "ai" else "🔍 Static Analysis"
                lines.append(f"*Source: {source_label}*")
                lines.append("")
                lines.append("---")
                lines.append("")

        return "\n".join(lines)


def _guess_lang(file_path: str) -> str:
    ext_map = {
        ".js": "javascript", ".jsx": "jsx", ".ts": "typescript", ".tsx": "tsx",
        ".py": "python", ".java": "java", ".css": "css", ".scss": "scss",
        ".html": "html", ".json": "json",
    }
    from pathlib import Path
    ext = Path(file_path).suffix.lower()
    return ext_map.get(ext, "")
